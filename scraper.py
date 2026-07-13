from __future__ import annotations

import argparse
import ast
import csv
import html
import json
import logging
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

BASE_URL = "https://izidream.bg/"
USER_AGENT = "Mozilla/5.0 (compatible; IzidreamTemuExporter/1.1; +https://izidream.bg/)"

# More-specific rules must be placed before broad rules.
CATEGORY_RULES = [
    (12044, "Home & Kitchen / Home Décor Products / Decorative Pillows / Throw Pillow Covers", ["декоративна калъф", "калъфка за декоратив"]),
    (12042, "Home & Kitchen / Home Décor Products / Decorative Pillows / Throw Pillows", ["декоративни възглав", "декоративна възглав"]),
    (11895, "Home & Kitchen / Bedding / Sheets & Pillowcases / Pillow Protectors", ["протектор за възглав"]),
    (11894, "Home & Kitchen / Bedding / Sheets & Pillowcases / Pillowcases", ["калъфки за възглав", "калъфка за възглав"]),
    (11915, "Home & Kitchen / Bedding / Bed Pillows & Positioners / Bed Pillows", ["възглавница за сън", "възглавници за спане", "анатомична възглав", "мемори възглав", "възглавниц"]),
    (12005, "Home & Kitchen / Bedding / Mattress Protectors & Encasements / Mattress Protectors", ["протектор за матрак", "протектори за матраци"]),
    (12002, "Home & Kitchen / Bedding / Mattress Pads & Toppers / Mattress Pads", ["подматрачна", "матрачен топер", "топ матрак"]),
    (11998, "Home & Kitchen / Bedding / Duvet Covers & Sets / Duvet Cover Sets", ["спален комплект", "спално бельо", "комплект спално"]),
    (11997, "Home & Kitchen / Bedding / Duvet Covers & Sets / Duvet Covers", ["плик за олекотена", "пликове за олекотени"]),
    (11896, "Home & Kitchen / Bedding / Sheets & Pillowcases / Fitted Sheets", ["чаршаф с ластик"]),
    (11897, "Home & Kitchen / Bedding / Sheets & Pillowcases / Flat Sheets", ["долен чаршаф", "чаршафи без ластик", "чаршаф без ластик"]),
    (12010, "Home & Kitchen / Bedding / Bedspreads, Coverlets & Sets / Bedspread & Coverlet Sets", ["комплект шалте", "комплект кувертюра"]),
    (12009, "Home & Kitchen / Bedding / Bedspreads, Coverlets & Sets / Bedspreads & Coverlets", ["шалте", "кувертюр", "покривало за легло"]),
    (53772, "Home & Kitchen / Home Décor Products / Slipcovers / Sofa Slipcovers / Sofa Throw Covers", ["покривало за диван", "шалте за диван"]),
    (12011, "Home & Kitchen / Bedding / Duvets & Down Comforters", ["олекотена завив", "зимни завив", "летни завив", "завивка"]),
    (11810, "Home & Kitchen / Bath / Towels / Towel Sets", ["комплект кърпи", "комплект хавли"]),
    (11812, "Home & Kitchen / Bath / Towels / Bath Towels", ["хавлиена кърпа", "кърпа за баня", "хавлия"]),
    (10492, "Home & Kitchen / Kitchen & Dining / Kitchen & Table Linens / Dish Cloths & Dish Towels", ["кухненска кърпа", "кухненски кърпи"]),
    (10498, "Home & Kitchen / Kitchen & Dining / Kitchen & Table Linens / Tablecloths", ["покривка за маса", "покривки за маса", "тишлайфер"]),
    (12906, "Home & Kitchen / Cleaning Supplies / Household Cleaners / Cleaning Tools / Cleaning Cloths", ["почистваща кърпа", "микрофибърна кърпа", "кърпа за почистване"]),
    (11901, "Home & Kitchen / Bedding / Blankets & Throws / Throws", ["одеяло за диван", "плед", "throw"]),
    (11899, "Home & Kitchen / Bedding / Blankets & Throws / Bed Blankets", ["одеяло", "поларено одеяло", "памучно одеяло"]),
    (26695, "Baby Products / Nursery / Bedding / Toddler Bedding / Bedding Sets", ["бебешки спален комплект", "комплект за кошара"]),
    (26696, "Baby Products / Nursery / Bedding / Toddler Bedding / Sheets, Pillowcases & Sets / Bed Sheets", ["бебешки чаршаф", "чаршаф за кошара"]),
    (11836, "Home & Kitchen / Bath / Kids' Bath / Kids' Bath Towels", ["детска хавлия", "детска кърпа"]),
    (29133, "Clothing, Shoes & Jewelry / Women / Clothing / Lingerie, Sleep & Lounge / Sleepwear / Robes", ["халат", "хавлиен халат"]),
]

PACKAGE_DIMENSIONS = {
    11915: (50.0, 35.0, 18.0), 12042: (45.0, 45.0, 15.0), 12044: (30.0, 25.0, 4.0),
    12011: (55.0, 40.0, 18.0), 11998: (38.0, 30.0, 10.0), 11997: (35.0, 27.0, 6.0),
    11896: (35.0, 27.0, 7.0), 11897: (35.0, 27.0, 7.0), 11894: (28.0, 22.0, 4.0),
    12009: (48.0, 38.0, 15.0), 12010: (50.0, 40.0, 16.0), 53772: (45.0, 35.0, 12.0),
    10498: (32.0, 25.0, 5.0), 10492: (25.0, 20.0, 4.0), 11812: (35.0, 28.0, 8.0),
    11810: (38.0, 30.0, 12.0), 11899: (45.0, 35.0, 14.0), 11901: (42.0, 32.0, 12.0),
    12005: (38.0, 30.0, 10.0), 12002: (55.0, 42.0, 20.0), 11895: (28.0, 22.0, 4.0),
    12906: (25.0, 20.0, 4.0), 26695: (38.0, 30.0, 10.0), 26696: (30.0, 24.0, 6.0),
    11836: (32.0, 25.0, 7.0), 29133: (38.0, 30.0, 8.0),
}
DEFAULT_CATEGORY = (11998, "Home & Kitchen / Bedding / Duvet Covers & Sets / Duvet Cover Sets")
DEFAULT_DIMS = (40.0, 30.0, 10.0)


@dataclass
class Product:
    url: str
    product_id: str
    sku: str
    name: str
    brand: str
    category_source: str
    category_id: int
    category_name: str
    description: str
    bullets: list[str]
    images: list[str]
    quantity: int
    price: float
    old_price: float | None
    weight_g: float
    length_cm: float
    width_cm: float
    height_cm: float
    country_origin: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    raw = html.unescape(str(value))
    if "<" in raw and ">" in raw:
        raw = BeautifulSoup(raw, "html.parser").get_text(" ", strip=True)
    return re.sub(r"\s+", " ", raw).strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip().casefold()


def session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "bg-BG,bg;q=0.9,en;q=0.7"})
    return s


def get(s: requests.Session, url: str, retries: int = 3) -> requests.Response:
    last = None
    for attempt in range(retries):
        try:
            r = s.get(url, timeout=35)
            r.raise_for_status()
            return r
        except requests.RequestException as exc:
            last = exc
            time.sleep(2 ** attempt)
    raise RuntimeError(f"Failed to download {url}: {last}")


def is_product_url(url: str) -> bool:
    p = urlparse(url)
    if p.netloc and "izidream.bg" not in p.netloc:
        return False
    path = p.path.rstrip("/")
    blocked = ("/blog", "/contacts", "/za-nas", "/user", "/cart", "/wishlist", "/search", "/promocii", "/podaruchni-vaucheri")
    if any(path.startswith(x) for x in blocked):
        return False
    return bool(re.search(r"-\d+$", path))


def discover_from_sitemaps(s: requests.Session) -> set[str]:
    candidates = [urljoin(BASE_URL, x) for x in ("sitemap.xml", "sitemap_index.xml", "sitemap-products.xml")]
    product_urls: set[str] = set()
    seen_maps: set[str] = set()

    def parse_map(url: str) -> None:
        if url in seen_maps:
            return
        seen_maps.add(url)
        try:
            text = get(s, url).text
        except Exception:
            return
        soup = BeautifulSoup(text, "xml")
        for loc in soup.find_all("loc"):
            target = loc.get_text(strip=True)
            if target.endswith(".xml"):
                parse_map(target)
            elif is_product_url(target):
                product_urls.add(target.split("#")[0])

    for candidate in candidates:
        parse_map(candidate)
    return product_urls


def discover_by_crawl(s: requests.Session, max_pages: int = 1800) -> set[str]:
    queue = [urljoin(BASE_URL, "produkti")]
    seen: set[str] = set()
    products: set[str] = set()
    allowed_hints = ("/produkti", "/p", "vazglavnici", "zaviv", "spalno", "charsh", "pokriv", "shalt", "havl", "odeyal", "protektor", "dete", "bebe", "tekstil")
    while queue and len(seen) < max_pages:
        url = queue.pop(0)
        if url in seen:
            continue
        seen.add(url)
        try:
            soup = BeautifulSoup(get(s, url).text, "html.parser")
        except Exception as exc:
            logging.warning("Skip discovery page %s: %s", url, exc)
            continue
        for a in soup.select("a[href]"):
            target = urljoin(url, a.get("href")).split("#")[0]
            if is_product_url(target):
                products.add(target)
            elif "izidream.bg" in urlparse(target).netloc and any(x in target for x in allowed_hints):
                if target not in seen and target not in queue:
                    queue.append(target)
        time.sleep(0.15)
    return products


def extract_event_json(page: str) -> dict[str, Any]:
    """Return the richest JSON payload embedded in data: JSON.parse('...')."""
    candidates: list[dict[str, Any]] = []
    pattern = re.compile(r"data:\s*JSON\.parse\('(.*?)'\)", re.DOTALL)
    for match in pattern.finditer(page):
        raw = match.group(1)
        try:
            decoded = ast.literal_eval("'" + raw.replace("'", "\\'") + "'")
            value = json.loads(decoded)
            if isinstance(value, dict):
                candidates.append(value)
        except Exception:
            continue
    if not candidates:
        return {}

    def score(data: dict[str, Any]) -> int:
        text = json.dumps(data, ensure_ascii=False)
        return sum(token in text for token in ("prod_code", "prod_name", "prod_body_short", "prod_weight", "oldprice", "prod_count"))

    return max(candidates, key=score)


def find_product_object(data: dict[str, Any]) -> dict[str, Any]:
    candidates = []
    if isinstance(data.get("object_data"), dict):
        candidates.append(data["object_data"])
    for key in ("themarketer", "object"):
        if isinstance(data.get(key), dict):
            candidates.append(data[key])
    candidates.sort(key=lambda x: sum(bool(x.get(k)) for k in ("prod_code", "prod_name", "prod_body_short", "prod_weight", "price")), reverse=True)
    for obj in candidates:
        if obj.get("prod_id") or obj.get("prod_code"):
            return obj
    return {}


def choose_category(name: str, source_cat: str, description: str) -> tuple[int, str]:
    """Choose the Temu category from the product itself, not incidental words in a set description."""
    primary = f"{name} {source_cat}".casefold()
    name_l = name.casefold()
    desc_l = description.casefold()

    # Strong product-type rules. The order matters: classify the actual item named
    # in the title before looking at component words inside the description.
    if any(x in name_l for x in ("комплект за кошара", "бебешки спален комплект", "спален комплект за беб")):
        return 26695, "Baby Products / Nursery / Bedding / Toddler Bedding / Bedding Sets"
    if "бебешки чаршаф" in name_l or "чаршаф за кошара" in name_l:
        return 26696, "Baby Products / Nursery / Bedding / Toddler Bedding / Sheets, Pillowcases & Sets / Bed Sheets"

    if "халат" in name_l:
        return 29133, "Clothing, Shoes & Jewelry / Women / Clothing / Lingerie, Sleep & Lounge / Sleepwear / Robes"
    if any(x in name_l for x in ("детска хавлия", "детска кърпа")):
        return 11836, "Home & Kitchen / Bath / Kids' Bath / Kids' Bath Towels"
    if any(x in name_l for x in ("кухненска кърпа", "кухненски кърпи")):
        return 10492, "Home & Kitchen / Kitchen & Dining / Kitchen & Table Linens / Dish Cloths & Dish Towels"
    if any(x in name_l for x in ("почистваща кърпа", "кърпа за почистване")) or ("микрофибърна кърпа" in name_l and "баня" not in name_l):
        return 12906, "Home & Kitchen / Cleaning Supplies / Household Cleaners / Cleaning Tools / Cleaning Cloths"
    if ("комплект" in name_l or "сет" in name_l) and any(x in name_l for x in ("кърп", "хавли")):
        return 11810, "Home & Kitchen / Bath / Towels / Towel Sets"
    if any(x in name_l for x in ("кърпа", "хавлия", "хавлиена кърпа")):
        return 11812, "Home & Kitchen / Bath / Towels / Bath Towels"

    if any(x in name_l for x in ("спален комплект", "спално бельо", "комплект спално")):
        return 11998, "Home & Kitchen / Bedding / Duvet Covers & Sets / Duvet Cover Sets"
    if "протектор за матрак" in name_l:
        return 12005, "Home & Kitchen / Bedding / Mattress Protectors & Encasements / Mattress Protectors"
    if "чаршаф с ластик" in name_l:
        return 11896, "Home & Kitchen / Bedding / Sheets & Pillowcases / Fitted Sheets"
    if "покривало за диван" in name_l or name_l.startswith("шалте за диван"):
        return 53772, "Home & Kitchen / Home Décor Products / Slipcovers / Sofa Slipcovers / Sofa Throw Covers"
    if any(x in name_l for x in ("шалте", "кувертюр", "покривало за легло")):
        return 12009, "Home & Kitchen / Bedding / Bedspreads, Coverlets & Sets / Bedspreads & Coverlets"
    if "декоратив" in name_l and "калъф" in name_l:
        return 12044, "Home & Kitchen / Home Décor Products / Decorative Pillows / Throw Pillow Covers"
    if "калъфка за декоратив" in name_l:
        return 12044, "Home & Kitchen / Home Décor Products / Decorative Pillows / Throw Pillow Covers"
    if any(x in name_l for x in ("декоративна възглав", "коледна възглав")) or ("градинск" in name_l and "възглав" in name_l):
        return 12042, "Home & Kitchen / Home Décor Products / Decorative Pillows / Throw Pillows"
    if "протектор за възглав" in name_l:
        return 11895, "Home & Kitchen / Bedding / Sheets & Pillowcases / Pillow Protectors"
    if any(x in name_l for x in ("калъфка за възглав", "калъфки за възглав")):
        return 11894, "Home & Kitchen / Bedding / Sheets & Pillowcases / Pillowcases"
    if any(x in name_l for x in ("долен чаршаф", "чаршаф без ластик", "чаршаф ранфорс")):
        return 11897, "Home & Kitchen / Bedding / Sheets & Pillowcases / Flat Sheets"
    if "плик за олекотена" in name_l:
        return 11997, "Home & Kitchen / Bedding / Duvet Covers & Sets / Duvet Covers"
    if any(x in name_l for x in ("олекотена завив", "лятна завив", "зимна завив")):
        return 12011, "Home & Kitchen / Bedding / Duvets & Down Comforters"
    if "одеяло" in name_l:
        return 11899, "Home & Kitchen / Bedding / Blankets & Throws / Bed Blankets"
    if "покривка за маса" in name_l or "тишлайфер" in name_l:
        return 10498, "Home & Kitchen / Kitchen & Dining / Kitchen & Table Linens / Tablecloths"
    if "възглав" in name_l:
        return 11915, "Home & Kitchen / Bedding / Bed Pillows & Positioners / Bed Pillows"

    # Then use the original ordered rules against name + breadcrumb category.
    for category_id, category_name, needles in CATEGORY_RULES:
        if any(n.casefold() in primary for n in needles):
            return category_id, category_name

    # Description is only a last-resort fallback.
    for category_id, category_name, needles in CATEGORY_RULES:
        if any(n.casefold() in desc_l for n in needles):
            return category_id, category_name

    logging.warning("No category rule matched: %s | %s. Using fallback %s", name, source_cat, DEFAULT_CATEGORY[0])
    return DEFAULT_CATEGORY

def extract_origin(description: str) -> str:
    match = re.search(r"страна\s+на\s+произход\s*[:\-]?\s*([А-ЯA-Z][А-Яа-яA-Za-z ]{2,30})", description, re.IGNORECASE)
    if match:
        value = match.group(1).strip().split(" Доставка")[0].strip(" .;,")
        normalized = {"българия": "Bulgaria", "турция": "Turkey", "китай": "China", "пакистан": "Pakistan", "индия": "India"}
        return normalized.get(value.casefold(), value)
    return "Bulgaria"


def parse_product_html(page: str, url: str) -> Product | None:
    soup = BeautifulSoup(page, "html.parser")

    ld_product: dict[str, Any] = {}
    for node in soup.select('script[type="application/ld+json"]'):
        try:
            value = json.loads(node.string or "{}")
        except Exception:
            continue
        if isinstance(value, dict) and value.get("@type") == "Product":
            ld_product = value
            break
    if not ld_product:
        logging.warning("Not a product page: %s", url)
        return None

    event = extract_event_json(page)
    obj = find_product_object(event)
    offer = ld_product.get("offers") if isinstance(ld_product.get("offers"), dict) else {}

    name = clean_text(obj.get("prod_name") or ld_product.get("name"))
    if not name or "ваучер" in name.casefold():
        return None

    id_match = re.search(r"-(\d+)$", urlparse(url).path.rstrip("/"))
    product_id = str(obj.get("prod_id") or (id_match.group(1) if id_match else ""))
    sku = str(obj.get("prod_code") or ld_product.get("sku") or product_id).strip()

    brand_value = ld_product.get("brand", {})
    brand = clean_text(obj.get("brand", {}).get("title") if isinstance(obj.get("brand"), dict) else "")
    if not brand and isinstance(brand_value, dict):
        brand = clean_text(brand_value.get("name"))
    brand = brand or "Izidream"

    category_source = ""
    if isinstance(event.get("category"), dict):
        category_source = clean_text(event["category"].get("hierarchy") or event["category"].get("title"))
    if not category_source and isinstance(obj.get("category"), dict):
        category_source = clean_text(obj["category"].get("hierarchy") or obj["category"].get("title"))
    if not category_source:
        crumbs = [clean_text(x) for x in soup.select(".breadcrumb__link")]
        category_source = " | ".join(crumbs[1:-1])

    description_html = obj.get("prod_body_short") or ""
    desc_soup = BeautifulSoup(html.unescape(str(description_html)), "html.parser")
    paragraph_bullets = [clean_text(node) for node in desc_soup.select("p, li")]
    paragraph_bullets = [x for x in paragraph_bullets if len(x) >= 8]
    description = clean_text(description_html) or clean_text(ld_product.get("description"))
    bullets = paragraph_bullets[:6]
    if not bullets:
        bullets = [x.strip() for x in re.split(r"(?<=[.!?])\s+|\s*\|\s*", description) if len(x.strip()) >= 15][:6]

    images: list[str] = []
    for el in soup.select(".gallery__thumb[data-zoom], .product-details__gallery-wrapper [data-zoom]"):
        src = el.get("data-zoom")
        if src:
            full = urljoin(BASE_URL, src)
            if full not in images:
                images.append(full)
    if not images and ld_product.get("image"):
        values = ld_product["image"] if isinstance(ld_product["image"], list) else [ld_product["image"]]
        images = [urljoin(BASE_URL, str(x)) for x in values]

    price = float(str(obj.get("price") or offer.get("price") or 0).replace(",", "."))
    old_price_raw = obj.get("oldprice") or obj.get("base_price")
    old_price = float(str(old_price_raw).replace(",", ".")) if old_price_raw and float(str(old_price_raw).replace(",", ".")) > price else None
    quantity = int(float(obj.get("prod_count") or (1 if "InStock" in str(offer.get("availability")) else 0)))

    weight_kg = float(str(obj.get("prod_weight") or 0).replace(",", "."))
    if not weight_kg:
        for prop in ld_product.get("additionalProperty", []) or []:
            if "тегло" in str(prop.get("name", "")).casefold():
                weight_kg = float(str(prop.get("value", "0")).replace(",", "."))
                break
    weight_g = max(1.0, round(weight_kg * 1000, 1))

    category_id, category_name = choose_category(name, category_source, description)
    length_cm, width_cm, height_cm = PACKAGE_DIMENSIONS.get(category_id, DEFAULT_DIMS)

    return Product(
        url=url,
        product_id=product_id,
        sku=sku,
        name=name,
        brand=brand,
        category_source=category_source,
        category_id=category_id,
        category_name=category_name,
        description=description[:2000],
        bullets=bullets,
        images=images[:95],
        quantity=max(0, quantity),
        price=round(price, 2),
        old_price=round(old_price, 2) if old_price else None,
        weight_g=weight_g,
        length_cm=length_cm,
        width_cm=width_cm,
        height_cm=height_cm,
        country_origin=extract_origin(description),
    )


def parse_product(s: requests.Session, url: str) -> Product | None:
    return parse_product_html(get(s, url).text, url)


def header_map(ws) -> dict[str, list[int]]:
    """Temu display headers are on row 2; internal field names are on row 4."""
    result: dict[str, list[int]] = {}
    for col in range(1, ws.max_column + 1):
        for value in (ws.cell(2, col).value, ws.cell(4, col).value):
            key = normalize_key(value)
            if key:
                result.setdefault(key, []).append(col)
    return result


def first_col(mapping: dict[str, list[int]], *keys: str) -> int | None:
    for key in keys:
        cols = mapping.get(normalize_key(key))
        if cols:
            return cols[0]
    return None


def all_cols(mapping: dict[str, list[int]], *keys: str) -> list[int]:
    out: list[int] = []
    for key in keys:
        out.extend(mapping.get(normalize_key(key), []))
    return sorted(set(out))


def set_value(ws, row: int, mapping: dict[str, list[int]], value: Any, *keys: str) -> None:
    col = first_col(mapping, *keys)
    if col is None:
        raise KeyError(f"Template column not found for any of: {keys}")
    ws.cell(row, col).value = value


def set_optional(ws, row: int, mapping: dict[str, list[int]], value: Any, *keys: str) -> bool:
    col = first_col(mapping, *keys)
    if col is None:
        return False
    ws.cell(row, col).value = value
    return True


def template_headers(ws) -> tuple[dict[int, str], dict[int, str]]:
    display = {col: clean_text(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)}
    technical = {col: clean_text(ws.cell(4, col).value) for col in range(1, ws.max_column + 1)}
    return display, technical


def read_dropdowns(book) -> dict[str, list[str]]:
    ws = book["Dropdown Lists"]
    result: dict[str, list[str]] = {}
    for values in ws.iter_rows(values_only=True):
        key = clean_text(values[0] if values else None)
        if not key:
            continue
        result[key] = [clean_text(value) for value in values[2:] if clean_text(value)]
    return result


def read_required_columns(book) -> dict[str, set[int]]:
    ws = book["GoodsLevelMode"]
    required: dict[str, set[int]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        key = clean_text(values[0] if values else None)
        if not key.endswith("_require"):
            continue
        category = key[:-8]
        required[category] = {
            col for col, value in enumerate(values, 1)
            if clean_text(value).casefold() == "require"
        }
    return required


DIMENSION_RE = re.compile(
    r"(?<!\d)(\d{1,3}(?:[.,]\d+)?)\s*(?:x|х|×|/|\*)\s*"
    r"(\d{1,3}(?:[.,]\d+)?)(?:\s*(?:x|х|×|/|\*)\s*(\d{1,3}(?:[.,]\d+)?))?\s*(?:см|cm)?",
    re.IGNORECASE,
)


def parse_number(value: str) -> float:
    return float(value.replace(",", "."))


def extract_dimension_tuples(text: str) -> list[tuple[float, float, float | None]]:
    values: list[tuple[float, float, float | None]] = []
    for match in DIMENSION_RE.finditer(text):
        a, b = parse_number(match.group(1)), parse_number(match.group(2))
        c = parse_number(match.group(3)) if match.group(3) else None
        # Ignore obvious years/phone fragments and implausibly tiny values.
        if 10 <= a <= 400 and 10 <= b <= 400:
            item = (a, b, c)
            if item not in values:
                values.append(item)
    return values


def labeled_dimension(text: str, labels: tuple[str, ...]) -> tuple[float, float, float | None] | None:
    label_pattern = "|".join(re.escape(x) for x in labels)
    pattern = re.compile(
        rf"(?:{label_pattern})[^0-9]{{0,55}}(\d{{1,3}}(?:[.,]\d+)?)\s*(?:x|х|×|/|\*)\s*"
        rf"(\d{{1,3}}(?:[.,]\d+)?)(?:\s*(?:x|х|×|/|\*)\s*(\d{{1,3}}(?:[.,]\d+)?))?",
        re.IGNORECASE,
    )
    match = pattern.search(text)
    if not match:
        return None
    a, b = parse_number(match.group(1)), parse_number(match.group(2))
    c = parse_number(match.group(3)) if match.group(3) else None
    return (a, b, c)


def choose_product_dimension(product: Product, label: str = "") -> tuple[float, float, float | None]:
    text = f"{product.name} {product.description}"
    label_l = label.casefold()
    special: tuple[str, ...] | None = None
    if "pillowcase" in label_l:
        special = ("калъфка", "калъфки")
    elif "duvet cover" in label_l:
        special = ("плик за завивка", "плик")
    elif "fitted sheet" in label_l or "flat sheet" in label_l:
        special = ("долен чаршаф", "чаршаф")
    elif "pillow" in label_l:
        special = ("размер", "възглавница")
    elif "duvet" in label_l or "comforter" in label_l:
        special = ("размер на завивката", "размер")
    if special:
        found = labeled_dimension(text, special)
        if found:
            return found
    name_dims = extract_dimension_tuples(product.name)
    if name_dims:
        return name_dims[0]
    desc_dims = extract_dimension_tuples(product.description)
    if desc_dims:
        return desc_dims[0]
    return (product.width_cm, product.length_cm, product.height_cm)


def extract_filling_weight(product: Product) -> float:
    text = f"{product.name} {product.description}"
    patterns = [
        r"(?:грамаж(?:\s+на\s+пълнежа)?|пълнеж)[^0-9]{0,35}(\d{2,5}(?:[.,]\d+)?)\s*(?:гр|g)",
        r"(\d{2,5}(?:[.,]\d+)?)\s*(?:гр|g)\s*(?:пълнеж|вата|пух|влакно)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return max(1.0, parse_number(match.group(1)))
    return max(50.0, round(product.weight_g * 0.8, 1))


COLOR_MAP = [
    (("многоцвет", "разноцвет", "шарен", "цветя"), "Multicolor"),
    (("тъмно син", "тъмносин", "индиго"), "Navy Blue"),
    (("светло син", "светлосин"), "Light Blue"),
    (("петрол",), "Teal"),
    (("тюркоаз",), "Turquoise"),
    (("аква",), "Aqua Blue"),
    (("циклама",), "Magenta"),
    (("бордо",), "Burgundy"),
    (("горчица",), "Mustard Yellow"),
    (("праскова",), "Peach"),
    (("капучино", "кафе"), "Coffee"),
    (("екрю", "крем"), "Creamy White"),
    (("бледо роз",), "Light Pink"),
    (("розов",), "Pink"),
    (("червен",), "Red"),
    (("оранжев",), "Orange"),
    (("жълт",), "Yellow"),
    (("зелен",), "Green"),
    (("лилав", "пърпъл"), "Purple"),
    (("син",), "Blue"),
    (("сив",), "Gray"),
    (("черен",), "Black"),
    (("кафяв",), "Brown"),
    (("бежов",), "Beige"),
    (("бял", "бяло", "бели"), "White"),
]


def extract_color(product: Product) -> str:
    text = f"{product.name} {product.description}".casefold()
    color_match = re.search(r"цвят\s*[:\-]?\s*([^.;]{1,80})", text, re.IGNORECASE)
    search_text = color_match.group(1) if color_match else text
    matches = []
    for needles, value in COLOR_MAP:
        if any(n in search_text for n in needles):
            matches.append(value)
    if len(set(matches)) > 1:
        return "Multicolor"
    return matches[0] if matches else "Multicolor"


MATERIAL_SYNONYMS = {
    "Cotton": ("памук", "ранфорс"),
    "Polyester": ("полиестер", "пе", "п.е", "микрофибър", "микросатен", "шерпа", "полар", "кадифе"),
    "Viscose": ("вискоза",),
    "Linen": ("лен", "ленен"),
    "Silk": ("коприна", "копринен"),
    "Wool": ("вълна", "вълнен"),
    "Nylon": ("найлон",),
    "Lyocell": ("лиосел",),
    "Polyurethane": ("полиуретан",),
    "bamboo": ("бамбук", "бамбуков"),
}


def _material_percent(text: str, synonyms: tuple[str, ...]) -> float | None:
    escaped = "|".join(re.escape(x) for x in synonyms)
    patterns = [
        rf"(\d{{1,3}}(?:[.,]\d+)?)\s*%\s*(?:[^.;,]{{0,22}})?(?:{escaped})",
        rf"(?:{escaped})(?:[^.;,]{{0,22}})?(\d{{1,3}}(?:[.,]\d+)?)\s*%",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return min(100.0, max(0.0, parse_number(match.group(1))))
    return None


def extract_materials(product: Product) -> dict[str, float]:
    text = f"{product.name} {product.description}".casefold()
    result: dict[str, float] = {}
    mentioned = []
    for material, synonyms in MATERIAL_SYNONYMS.items():
        if any(s in text for s in synonyms):
            mentioned.append(material)
            percent = _material_percent(text, synonyms)
            if percent is not None:
                result[material] = percent
    if result:
        total = sum(result.values())
        if total > 100.5:
            result = {k: round(v * 100 / total, 1) for k, v in result.items()}
        elif total < 99.5 and len(result) == 1:
            only = next(iter(result))
            result[only] = 100.0
        return result
    if len(mentioned) == 1:
        return {mentioned[0]: 100.0}
    if len(mentioned) > 1:
        share = round(100 / len(mentioned), 1)
        return {m: share for m in mentioned}
    return {"Polyester": 100.0}


FILLING_SYNONYMS = [
    (("мемори",), "Memory foam"),
    (("полиуретанова пяна", "полиуретаново ядро", "полиуретанова"), "Polyurethane Foam"),
    (("латекс",), "Latex"),
    (("вълна", "вълнен пълнеж"), "Wool"),
    (("памучен пълнеж",), "Cotton"),
    (("силиконов пух", "силиконова вата", "силиконово влакно", "силиконов пълнеж"), "Silicone Fiber"),
    (("пух", "пера"), "Down"),
    (("полиестер", "пе", "вата", "рециклируем материал"), "Polyester"),
]


def extract_filling_materials(product: Product) -> dict[str, float]:
    text = f"{product.name} {product.description}".casefold()
    for needles, material in FILLING_SYNONYMS:
        if any(n in text for n in needles):
            return {material: 100.0}
    return {"Polyester": 100.0}


def closest_allowed(value: str, allowed: list[str], default: str | None = None) -> str:
    if value in allowed:
        return value
    folded = {x.casefold(): x for x in allowed}
    if value.casefold() in folded:
        return folded[value.casefold()]
    return default or (allowed[0] if allowed else value)


def normalized_size(a: float, b: float) -> str:
    def n(x: float) -> str:
        return str(int(x)) if float(x).is_integer() else str(round(x, 1))
    return f"{n(a)}cm*{n(b)}cm"


def choose_size_fields(category: int, product: Product, dropdowns: dict[str, list[str]], field_name: str = "Size") -> tuple[str, str, str]:
    a, b, _ = choose_product_dimension(product, field_name)
    candidate = normalized_size(a, b)
    family_values = dropdowns.get(f"t_4_{category}_Size Family", [])
    regular = "2 - Regular Size" if "2 - Regular Size" in family_values else (family_values[0] if family_values else "2 - Regular Size")
    regular_sub_values = dropdowns.get(f"t_4_{category}_{regular}_Sub-Size Family", [])
    sub = "10 - Alpha" if "10 - Alpha" in regular_sub_values else (regular_sub_values[0] if regular_sub_values else "10 - Alpha")
    size_values = dropdowns.get(f"t_4_{category}_{regular}_{sub}_{field_name}", [])
    if candidate in size_values:
        return regular, sub, candidate
    reversed_candidate = normalized_size(b, a)
    if reversed_candidate in size_values:
        return regular, sub, reversed_candidate
    custom = "101 - Custom size" if "101 - Custom size" in family_values else regular
    custom_sub_values = dropdowns.get(f"t_4_{category}_{custom}_Sub-Size Family", [])
    custom_sub = "10 - Alpha" if "10 - Alpha" in custom_sub_values else (custom_sub_values[0] if custom_sub_values else sub)
    return custom, custom_sub, candidate


def primary_material_name(product: Product, allowed: list[str] | None = None) -> str:
    materials = extract_materials(product)
    primary = max(materials, key=materials.get)
    aliases = {
        "bamboo": "Viscose",
        "Polyester": "Polyester",
        "Cotton": "Cotton",
        "Linen": "Linen",
        "Silk": "Silk",
        "Wool": "Wool",
        "Nylon": "Nylon",
        "Viscose": "Viscose",
        "Lyocell": "Lyocell",
        "Polyurethane": "Polyurethane",
    }
    value = aliases.get(primary, primary)
    if allowed:
        for candidate in (value, "Microfiber" if "микрофибър" in product.description.casefold() else "", "Polyester", "Cotton"):
            if candidate and candidate in allowed:
                return candidate
        return allowed[0]
    return value


def property_group(technical_header: str) -> tuple[str, str] | None:
    match = re.match(r"t_3_Property:(\d+):(.+)$", technical_header)
    if match:
        return match.group(1), match.group(2)
    return None


def fill_option_group(ws, row: int, cols: list[int], display: dict[int, str], values: dict[str, float]) -> None:
    available = {display[col].split(":", 1)[1].casefold(): col for col in cols if ":" in display[col]}
    written = False
    for option, percentage in values.items():
        aliases = [option]
        if option == "bamboo":
            aliases += ["Bamboo", "Viscose", "Other Fibers", "Other Materials"]
        if option == "Polyurethane Foam":
            aliases += ["Polyurethane", "Foam"]
        if option == "Silicone Fiber":
            aliases += ["Silicone", "Polyester"]
        for alias in aliases:
            col = available.get(alias.casefold())
            if col:
                ws.cell(row, col).value = round(percentage, 1)
                written = True
                break
    if not written and cols:
        # Prefer common textile fallbacks instead of an arbitrary exotic option.
        for fallback in ("polyester", "cotton", "n/a"):
            if fallback in available:
                ws.cell(row, available[fallback]).value = 100
                return
        ws.cell(row, cols[0]).value = 100


def fill_required_attributes(
    ws,
    row: int,
    product: Product,
    mapping: dict[str, list[int]],
    display: dict[int, str],
    technical: dict[int, str],
    dropdowns: dict[str, list[str]],
    required_cols: set[int],
) -> None:
    category = product.category_id
    text = f"{product.name} {product.description}"

    # Required option groups such as cover material, filling material and composition.
    grouped: dict[str, list[int]] = {}
    for col in required_cols:
        group = property_group(technical.get(col, ""))
        if group:
            grouped.setdefault(group[0], []).append(col)
    for property_id, cols in grouped.items():
        if any(ws.cell(row, col).value not in (None, "") for col in cols):
            continue
        label = display[cols[0]].split(":", 1)[0].casefold()
        if property_id == "2018" or "filling material" in label:
            fill_option_group(ws, row, cols, display, extract_filling_materials(product))
        else:
            fill_option_group(ws, row, cols, display, extract_materials(product))

    # Single-value product attributes required for selected categories.
    for col in sorted(required_cols):
        tech = technical.get(col, "")
        label = display.get(col, "")
        if not tech.startswith("t_3_Property:") or property_group(tech):
            continue
        if ws.cell(row, col).value not in (None, ""):
            continue
        key = f"t_3_{category}_{label}"
        allowed = dropdowns.get(key, [])
        label_l = label.casefold()
        if "power supply" in label_l:
            value = closest_allowed("Use Without Electricity", allowed, "Use Without Electricity")
        elif "battery properties" in label_l:
            value = closest_allowed("Without Battery", allowed, "Without Battery")
        elif "material type" in label_l:
            value = closest_allowed("Textile Material", allowed, "Textile Material")
        elif "material" in label_l and "square" not in label_l:
            value = primary_material_name(product, allowed)
        elif "square gram weight" in label_l:
            value = closest_allowed("150-160g", allowed, allowed[0] if allowed else "150-160g")
        elif "fibre composition" in label_l:
            value = "Yes"
        elif "applicable age" in label_l:
            preferred = "Infant" if any(x in text.casefold() for x in ("беб", "кошара", "детск")) else "Adult"
            match = next((x for x in allowed if preferred.casefold() in x.casefold()), None)
            value = match or (allowed[0] if allowed else preferred)
        else:
            value = allowed[0] if allowed else "Not Applicable"
        ws.cell(row, col).value = value

    # Variation fields.
    theme_col = first_col(mapping, "t_4_Variation Theme")
    if theme_col and theme_col in required_cols:
        themes = dropdowns.get(f"t_4_{category}_Variation Theme", [])
        structured_size_col = first_col(mapping, "t_4_Size:50546518", "t_4_Size:3001")
        has_structured = any(
            first_col(mapping, key) in required_cols
            for key in ("t_4_Size Family", "t_4_Sub-Size Family", "t_4_Size:50546518", "t_4_Size:3001", "t_4_Sale Property:1001")
        )
        if has_structured:
            preferred = themes[0] if themes else "Color × Size"
        else:
            preferred = "Color × Size" if "Color × Size" in themes else ("Size" if "Size" in themes else (themes[0] if themes else "Color × Size"))
        ws.cell(row, theme_col).value = preferred

        if has_structured:
            size_field = "Duvet Cover Size" if first_col(mapping, "t_4_Size:50546518") in required_cols else "Size"
            family, sub, size = choose_size_fields(category, product, dropdowns, size_field)
            set_optional(ws, row, mapping, family, "t_4_Size Family")
            set_optional(ws, row, mapping, sub, "t_4_Sub-Size Family")
            if size_field == "Duvet Cover Size":
                set_optional(ws, row, mapping, size, "t_4_Size:50546518")
            else:
                set_optional(ws, row, mapping, size, "t_4_Size:3001")
            allowed_colors = dropdowns.get(f"t_4_{category}_Color", [])
            color = closest_allowed(extract_color(product), allowed_colors, "Multicolor" if "Multicolor" in allowed_colors else (allowed_colors[0] if allowed_colors else extract_color(product)))
            set_optional(ws, row, mapping, color, "t_4_Sale Property:1001")
        else:
            a, b, _ = choose_product_dimension(product)
            if "Color" in preferred:
                set_optional(ws, row, mapping, extract_color(product), "t_4_Custom Spec:1001")
            if "Size" in preferred:
                set_optional(ws, row, mapping, normalized_size(a, b), "t_4_Custom Spec:3001")

    # Unit used by the size chart.
    unit_col = first_col(mapping, "t_5_Unit")
    if unit_col and unit_col in required_cols:
        ws.cell(row, unit_col).value = "cm-g-ml"

    # Product dimensions in the category-specific size chart.
    for col in sorted(required_cols):
        tech = technical.get(col, "")
        if not tech.startswith("t_5_Size Chart Element:"):
            continue
        label = display.get(col, "")
        a, b, c = choose_product_dimension(product, label)
        label_l = label.casefold()
        if "filling weight" in label_l:
            value = extract_filling_weight(product)
        elif "width" in label_l:
            value = a
        elif "length" in label_l:
            value = b
        elif "height" in label_l:
            value = c or (20.0 if "sheet" in label_l else 12.0)
        elif "chest" in label_l:
            value = max(a, b, 100.0)
        else:
            value = a
        ws.cell(row, col).value = round(float(value), 1)

    # Packaging and offer fields that are marked required in the template.
    if first_col(mapping, "t_6_Individually packed") in required_cols:
        set_optional(ws, row, mapping, "Yes", "t_6_Individually packed")
    if first_col(mapping, "t_6_Total packaging quantity") in required_cols:
        set_optional(ws, row, mapping, 1, "t_6_Total packaging quantity")
    if first_col(mapping, "t_6_Packaging unit") in required_cols:
        unit = "pack" if any(x in product.name.casefold() for x in ("комплект", "сет")) else "piece"
        set_optional(ws, row, mapping, unit, "t_6_Packaging unit")
    if first_col(mapping, "t_7_Shipping Template") in required_cols:
        shipping_values = dropdowns.get("t_7_Shipping Template", [])
        set_optional(ws, row, mapping, shipping_values[0] if shipping_values else "Bulgaria", "t_7_Shipping Template")

    # These fields are required by Temu's data definitions even when the red rule is not category-specific.
    fulfillment_values = dropdowns.get("t_7_Fulfillment Channel", [])
    set_optional(ws, row, mapping, fulfillment_values[0] if fulfillment_values else "I will ship this item myself", "t_7_Fulfillment Channel")


def fill_template(template: Path, output: Path, products: list[Product]) -> None:
    book = load_workbook(template)
    ws = book["Template"]
    mapping = header_map(ws)
    display, technical = template_headers(ws)
    dropdowns = read_dropdowns(book)
    required_by_category = read_required_columns(book)
    start_row = 5

    if normalize_key(ws.cell(4, 20).value) != normalize_key("t_2_Product Description"):
        raise RuntimeError("Unexpected Temu template structure: technical row 4 is missing or modified.")

    for offset, p in enumerate(products):
        row = start_row + offset
        set_value(ws, row, mapping, p.category_id, "t_1_Category", "Category")
        set_value(ws, row, mapping, p.category_name, "t_1_Category Name", "Category Name")
        set_value(ws, row, mapping, p.name, "t_1_Product Name", "Product Name")
        set_value(ws, row, mapping, p.sku, "t_1_Contribution Goods", "Contribution Goods")
        set_value(ws, row, mapping, p.sku, "t_1_Contribution SKU", "Contribution SKU")
        set_value(ws, row, mapping, "Add", "t_1_Update or Add", "Update or Add")
        set_value(ws, row, mapping, p.brand, "t_1_Brand", "Brand")
        set_value(ws, row, mapping, p.brand, "t_1_Trademark", "Trademark")
        set_value(ws, row, mapping, p.description, "t_2_Product Description", "Product Description")

        bullet_cols = all_cols(mapping, "t_2_Bullet Point", "Bullet Point")
        for i, col in enumerate(bullet_cols[:6]):
            ws.cell(row, col).value = p.bullets[i] if i < len(p.bullets) else None

        detail_cols = all_cols(mapping, "t_2_Detail Images URL", "Detail Images URL")
        for i, col in enumerate(detail_cols):
            ws.cell(row, col).value = p.images[i] if i < len(p.images) else None

        sku_image_cols = all_cols(mapping, "t_6_SKU Images URL", "SKU Images URL")
        if sku_image_cols:
            ws.cell(row, sku_image_cols[0]).value = p.images[0] if p.images else ""

        set_value(ws, row, mapping, p.quantity, "t_6_Quantity")
        set_value(ws, row, mapping, p.price, "t_6_Base Price - EUR", "Base Price - EUR")
        set_value(ws, row, mapping, p.url, "t_6_Reference Link", "Reference Link")
        if p.old_price:
            set_value(ws, row, mapping, p.old_price, "t_6_List Price - EUR", "List Price - EUR")
        else:
            set_value(ws, row, mapping, "N/A", "t_6_Not available for List price", "Not available for List price")
        set_value(ws, row, mapping, p.weight_g, "t_6_Weight - g", "Weight - g")
        set_value(ws, row, mapping, p.length_cm, "t_6_Length - cm", "Length - cm")
        set_value(ws, row, mapping, p.width_cm, "t_6_Width - cm", "Width - cm")
        set_value(ws, row, mapping, p.height_cm, "t_6_Height - cm", "Height - cm")
        set_value(ws, row, mapping, "1 Day", "t_7_Handling Time", "Handling Time")
        set_value(ws, row, mapping, "GEN STANDARD", "t_7_Item Tax Code", "Item Tax Code")
        set_value(ws, row, mapping, p.country_origin, "t_8_Country/Region of Origin", "Country/Region of Origin")

        required_cols = required_by_category.get(str(p.category_id), set())
        fill_required_attributes(ws, row, p, mapping, display, technical, dropdowns, required_cols)

    book.save(output)
    validate_output(output, len(products))


def validate_output(path: Path, expected_rows: int) -> None:
    book = load_workbook(path, read_only=False, data_only=False)
    ws = book["Template"]
    mapping = header_map(ws)
    display, technical = template_headers(ws)
    required_by_category = read_required_columns(book)
    start_row = 5

    if normalize_key(ws.cell(4, 20).value) != normalize_key("t_2_Product Description"):
        raise RuntimeError("Validation failed: Temu technical row 4 was overwritten.")

    errors: list[str] = []
    for row in range(start_row, start_row + expected_rows):
        category = clean_text(ws.cell(row, first_col(mapping, "t_1_Category")).value)
        required_cols = required_by_category.get(category, set())
        option_groups: dict[str, list[int]] = {}
        simple_cols: list[int] = []
        for col in required_cols:
            group = property_group(technical.get(col, ""))
            if group:
                option_groups.setdefault(group[0], []).append(col)
            else:
                simple_cols.append(col)
        for property_id, cols in option_groups.items():
            if not any(ws.cell(row, col).value not in (None, "") for col in cols):
                errors.append(f"row {row}: required attribute group {display[cols[0]].split(':', 1)[0]} is blank")
        for col in simple_cols:
            value = ws.cell(row, col).value
            # List price may be replaced by the explicit N/A field.
            if technical.get(col) == "t_6_List Price - EUR":
                na_col = first_col(mapping, "t_6_Not available for List price")
                if value in (None, "") and (na_col is None or ws.cell(row, na_col).value in (None, "")):
                    errors.append(f"row {row}: List Price and N/A are both blank")
            elif value in (None, ""):
                errors.append(f"row {row}: required field {display.get(col) or technical.get(col)} is blank")
        # Validate conditional custom variation fields for generic categories.
        theme_col = first_col(mapping, "t_4_Variation Theme")
        theme = clean_text(ws.cell(row, theme_col).value) if theme_col else ""
        if "Color" in theme:
            structured_color = first_col(mapping, "t_4_Sale Property:1001")
            custom_color = first_col(mapping, "t_4_Custom Spec:1001")
            if all(col is None or ws.cell(row, col).value in (None, "") for col in (structured_color, custom_color)):
                errors.append(f"row {row}: variation color is blank")
        if "Size" in theme:
            size_cols = [
                first_col(mapping, "t_4_Size:50546518"),
                first_col(mapping, "t_4_Size:3001"),
                first_col(mapping, "t_4_Custom Spec:3001"),
            ]
            if all(col is None or ws.cell(row, col).value in (None, "") for col in size_cols):
                errors.append(f"row {row}: variation size is blank")
        if len(errors) >= 40:
            break
    if errors:
        raise RuntimeError("Validation failed:\n" + "\n".join(errors))
    logging.info("Output validation passed: %d product rows and all category-required groups populated", expected_rows)

def export_raw(path: Path, products: list[Product]) -> None:
    fields = list(asdict(products[0]).keys()) if products else ["url"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for p in products:
            row = asdict(p)
            row["bullets"] = " | ".join(p.bullets)
            row["images"] = " | ".join(p.images)
            writer.writerow(row)


def main() -> None:
    parser = argparse.ArgumentParser(description="Scrape Izidream products and fill the supplied Temu template.")
    parser.add_argument("--template", default="template.xlsx")
    parser.add_argument("--output", default="TEMU_IZIDREAM_UPLOAD.xlsx")
    parser.add_argument("--raw", default="izidream_raw_export.csv")
    parser.add_argument("--limit", type=int, default=20, help="0 means all products")
    parser.add_argument("--urls-file", help="Optional text file containing one product URL per line")
    parser.add_argument("--delay", type=float, default=0.35)
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    s = session()
    if args.urls_file:
        urls = {x.strip() for x in Path(args.urls_file).read_text(encoding="utf-8").splitlines() if x.strip()}
    else:
        urls = discover_from_sitemaps(s)
        if not urls:
            logging.info("No usable sitemap found; using category crawl fallback")
            urls = discover_by_crawl(s)
    urls = sorted(urls)
    if args.limit > 0:
        urls = urls[:args.limit]
    logging.info("Product URLs selected: %d", len(urls))

    products: list[Product] = []
    for index, url in enumerate(urls, 1):
        try:
            product = parse_product(s, url)
            if product:
                products.append(product)
                logging.info("[%d/%d] SKU=%s | %s | Category=%s", index, len(urls), product.sku, product.name, product.category_id)
        except Exception as exc:
            logging.exception("Failed product %s: %s", url, exc)
        time.sleep(max(0, args.delay))

    if not products:
        raise RuntimeError("No product data was extracted. Output file was not created.")

    fill_template(Path(args.template), Path(args.output), products)
    export_raw(Path(args.raw), products)
    logging.info("Done: %s products -> %s", len(products), args.output)


if __name__ == "__main__":
    main()
